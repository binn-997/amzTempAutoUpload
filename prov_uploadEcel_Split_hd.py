import pandas as pd
import re
import math
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# ======================== 1. 统一配置层 ========================
class Config:
    SOURCE_FILE = "610hd.xlsx"
    TEMPLATE_FILE = "./temp/prov_hdTemp.xlsx"
    OUTPUT_SUFFIX = "_prov"

    # ========== ★ 拆分表配置 ★ ==========
    CHUNK_SIZE = 1000                 # 每个拆分文件最多包含多少行数据
    OUTPUT_DIR = "./prov_output"          # 拆分文件保存目录（None 则保存在当前目录）

    # ========== ★ 标题前缀（Item Name 品牌名）★ ==========
    TITLE_PREFIX = ""

    MAX_SOURCE_COLS = 60
    START_ROW = 4
    MAX_IMAGES = 7

    # ========== 源数据列索引 (0-based) ==========
    SRC = {
        "sku": 1,
        "parent_sku": 2,
        "title": 3,
        "color": 6,
        "size": 7,
        "package_weight": 9,
        "standprice": 38,
        "list_price_tax": 39,
        "generic_keywords": 45,
        "bullet1": 40,
        "bullet2": 41,
        "bullet3": 42,
        "bullet4": 43,
        "bullet5": 44,
        "images": [10, 11, 12, 13, 14, 15, 16],
    }
    # ========== ★ 模块化映射配置（修改映射规则只需改这里）★ ==========

    # --- A. 一对一映射：{DataFrame列名: 目标Excel列字母} ---
    SIMPLE_MAP = {
        "sku":            "B",
        "parent_sku":     "BY",
        "title":          "G",
        "standprice":     "V",
        "package_weight": "GD",
        "list_price_tax": "KP",
        "final_keywords": "CE",
        "size_class":     "BA",
    }

    # --- B. 一对多映射：{DataFrame列名: [目标Excel列字母列表]} ---
    MULTI_MAP = {
        "color": ["CF", "CG"],
        "size":  ["BB", "CH", "FC"],
    }

    # --- C. Bullet Points 映射 ---
    BULLET_SRC_PREFIX = "final_bullet"
    BULLET_TGT_COLS   = ["CO", "CP", "CQ", "CR", "CS"]

    # --- D. 图片映射 ---
    IMG_SRC_COLS  = [10, 11, 12, 13, 14, 15, 16]
    IMG_TGT_START = "BH"

    # --- E. 父子标识映射 ---
    PC_COL = "BX"

    # ========== 父体清除规则 ==========
    PARENT_CLEAR_LETTERS = [
        "V", "W", "X", "AA", "AB", "BA", "BB", "BC", "BF", "BG", "AZ",
        "BY", "BZ", "CF", "CG", "CH", "FC",
    ]
    PARENT_CLEAR_IDX = [column_index_from_string(c) for c in PARENT_CLEAR_LETTERS]

    # ========== 尺码替换规则 ==========
    SIZE_REPLACEMENTS = {
        r"\bXXXL\b":        "3XL",
        r"\bXXXXL\b":       "4XL",
        r"\bXXXXXL\b":      "5XL",
        r"\bXXXXXXL\b":     "6XL",
        r"\bXXXXXXXL\b":    "7XL",
        r"\bXXXXXXXXL\b":   "8XL",
        r"\bXXXXXXXXXL\b":  "9XL",
        r"\bone size\b":    "Einheitsgröße",
    }


# ======================== 2. 数据预处理层 ========================
def preprocess_source_data(df: pd.DataFrame) -> pd.DataFrame:
    """在内存中完成所有数据清洗与父子体继承"""
    print("预处理源数据：清洗空值、替换尺码、继承父体数据...")

    if df.shape[1] < Config.MAX_SOURCE_COLS:
        for i in range(df.shape[1], Config.MAX_SOURCE_COLS):
            df[i] = ''
    df = df.iloc[:, :Config.MAX_SOURCE_COLS].fillna('')
    df = df.apply(lambda col: col.str.strip() if col.dtype == 'object' else col)

    rename_map = {v: k for k, v in Config.SRC.items() if isinstance(v, int)}
    df = df.rename(columns=rename_map)

    # ★ Item Name 加品牌前缀
    df['title'] = Config.TITLE_PREFIX + df['title'].astype(str)

    for pattern, replacement in Config.SIZE_REPLACEMENTS.items():
        df['size'] = df['size'].astype(str).str.replace(
            pattern, replacement, regex=True, flags=re.IGNORECASE
        )

    df['size_class'] = df['size'].apply(
        lambda x: "Numerisch" if str(x).strip().isdigit() else "Alphanumerisch"
    )

    df['is_parent'] = df['sku'] == df['parent_sku']

    parent_cols = ['generic_keywords', 'bullet1', 'bullet2', 'bullet3', 'bullet4', 'bullet5']
    parent_df = df[df['is_parent']][['sku'] + parent_cols].set_index('sku')
    parent_dict = parent_df.to_dict(orient='index')

    def get_parent_data(row):
        if row['is_parent']:
            return pd.Series([row[col] for col in parent_cols])
        p_data = parent_dict.get(row['parent_sku'], {})
        return pd.Series([p_data.get(col, "") for col in parent_cols])

    df[['final_keywords'] + [f'final_bullet{i}' for i in range(1, 6)]] = \
        df.apply(get_parent_data, axis=1)

    return df


# ======================== 模块一：数据填充（数据驱动版） ========================
def fill_data_to_template(ws, df: pd.DataFrame):
    """根据 Config 中的映射配置，将 DataFrame 逐行写入 Excel"""

    simple_idx = {k: column_index_from_string(v) for k, v in Config.SIMPLE_MAP.items()}
    multi_idx  = {k: [column_index_from_string(c) for c in v]
                  for k, v in Config.MULTI_MAP.items()}
    bullet_idx = [column_index_from_string(c) for c in Config.BULLET_TGT_COLS]
    img_start  = column_index_from_string(Config.IMG_TGT_START)
    pc_idx     = column_index_from_string(Config.PC_COL)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=Config.START_ROW):

        # ---- A. 一对一映射 ----
        for src_col, tgt_idx in simple_idx.items():
            ws.cell(row=r_idx, column=tgt_idx, value=row[src_col])

        # ---- B. 一对多映射 ----
        for src_col, tgt_indices in multi_idx.items():
            for tgt_idx in tgt_indices:
                ws.cell(row=r_idx, column=tgt_idx, value=row[src_col])

        # ---- C. Bullet Points ----
        for i, tgt_idx in enumerate(bullet_idx):
            src_col = f"{Config.BULLET_SRC_PREFIX}{i + 1}"
            ws.cell(row=r_idx, column=tgt_idx, value=row[src_col])

        # ---- D. 图片 ----
        images = [str(row[pos]).strip() for pos in Config.IMG_SRC_COLS
                  if str(row[pos]).strip()]
        for img_i, url in enumerate(images[:Config.MAX_IMAGES]):
            ws.cell(row=r_idx, column=img_start + img_i, value=url)
        for empty_i in range(len(images), Config.MAX_IMAGES):
            ws.cell(row=r_idx, column=img_start + empty_i, value="")

        # ---- E. 父子标识 ----
        pc_val = "Parent" if row['is_parent'] else "Child"
        ws.cell(row=r_idx, column=pc_idx, value=pc_val)


# ======================== 模块二：清除父体数据 ========================
def clear_parent_rows(ws, df: pd.DataFrame):
    """独立清除模块：抹平父体行的子体专属字段"""
    # 将全局索引映射到局部 df 的行号
    parent_local_indices = df[df['is_parent']].index.tolist()
    if not parent_local_indices:
        return

    pc_idx = column_index_from_string(Config.PC_COL)

    # 局部 df 的第 N 行对应模板的第 Config.START_ROW + N 行 (0-based)
    for local_row_idx in parent_local_indices:
        excel_row = Config.START_ROW + local_row_idx

        merged_to_remove = [
            mr for mr in ws.merged_cells.ranges
            if mr.min_row <= excel_row <= mr.max_row
        ]
        for mr in merged_to_remove:
            ws.unmerge_cells(str(mr))

        for col_idx in Config.PARENT_CLEAR_IDX:
            ws.cell(row=excel_row, column=col_idx, value="")

        ws.cell(row=excel_row, column=pc_idx, value="Parent")


# ======================== 模块三：拆分并保存 ========================
def save_split_workbooks(processed_df: pd.DataFrame):
    """
    将数据按 CHUNK_SIZE 拆分为多个文件。
    每个文件独立加载模板 → 填充 → 清除父体 → 删除多余行 → 保存。
    """
    total_rows = len(processed_df)
    if total_rows == 0:
        print("源数据为空，跳过。")
        return

    # ★ 准备输出目录
    output_dir = Config.OUTPUT_DIR
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"📁 创建输出目录: {output_dir}")

    base_name = os.path.splitext(os.path.basename(Config.SOURCE_FILE))[0]
    num_chunks = math.ceil(total_rows / Config.CHUNK_SIZE)

    print(f"\n数据预处理完成，共 {total_rows} 行。")
    print(f"按每文件 {Config.CHUNK_SIZE} 行拆分，共 {num_chunks} 个文件。\n")

    for chunk_idx in range(num_chunks):
        start = chunk_idx * Config.CHUNK_SIZE
        end   = min(start + Config.CHUNK_SIZE, total_rows)

        # 关键：用 .iloc[start:end].reset_index(drop=True) 得到 0-based 的局部索引
        chunk_df = processed_df.iloc[start:end].reset_index(drop=True)
        chunk_rows = len(chunk_df)

        # 加载模板
        wb = load_workbook(Config.TEMPLATE_FILE, keep_vba=False)
        if 'Vorlage' not in wb.sheetnames:
            print("错误：模板中未找到 Vorlage 工作表")
            exit(1)
        ws = wb['Vorlage']

        # 填充 & 清除
        fill_data_to_template(ws, chunk_df)
        clear_parent_rows(ws, chunk_df)

        # 删除多余空行
        last_used_row = Config.START_ROW + chunk_rows - 1
        start_delete_row = last_used_row + 1
        if ws.max_row >= start_delete_row:
            ws.delete_rows(start_delete_row, amount=ws.max_row - start_delete_row + 1)

        # ★ 命名：{源文件名}_esfa_part{序号}.xlsx
        output_name = f"{base_name}{Config.OUTPUT_SUFFIX}_part{chunk_idx + 1}.xlsx"
        if output_dir:
            output_path = os.path.join(output_dir, output_name)
        else:
            output_path = output_name

        print(f"  [{chunk_idx + 1}/{num_chunks}] 保存: {output_path}  ({chunk_rows} 行)")
        wb.save(output_path)
        wb.close()

    print(f"\n✅ 全部 {num_chunks} 个文件保存完毕。")


# ======================== 主程序调度 ========================
def main():
    print("🚀 开始执行亚马逊模板处理程序...")

    print("\n[1/3] 读取源数据...")
    raw_df = pd.read_excel(Config.SOURCE_FILE, sheet_name=0, header=None, dtype=str)

    print("\n[2/3] 预处理源数据...")
    processed_df = preprocess_source_data(raw_df)

    print("\n[3/3] 执行拆分写入...")
    save_split_workbooks(processed_df)

    print("\n🎉 全部流程执行完毕！")


if __name__ == "__main__":
    main()
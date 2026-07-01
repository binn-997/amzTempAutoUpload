"""
统一 CLI 入口 — 配置驱动的品牌处理器 + 利润定价工具。

用法:
    # 推荐：作为模块运行
    python -m src.main process --type azd

    # 也可直接运行脚本
    python src/main.py process --type azd

    # 覆盖源文件和分块大小
    python src/main.py process --type azd --source 624azh.xlsx --chunk 800

    # 使用自定义配置文件
    python src/main.py process --type hd --config my_config.yaml

    # 仅预览，不写入文件
    python src/main.py process --type th --dry-run

    # 列出所有可用产品类型
    python src/main.py list-types
    python src/main.py list-types --config my_config.yaml

    # ★ 利润定价：根据成本+重量反算销售价并回填到源文件 AM 列
    python -m src.main price --source data/630td.xlsx
    python -m src.main price --source data/630td.xlsx --target 0.25 --rate 7.8
    python -m src.main price --source data/630td.xlsx --cost-col I --weight-col J

    # ★ 颜色翻译：英→德翻译 + 变体 V1/V2 标注，原地修改源文件 G/H 列
    python -m src.main translate --source data/701th.xlsx
    python -m src.main translate --source data/xxx.xlsx --no-translate
    python -m src.main translate --source data/xxx.xlsx --color-col E --size-col F
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 修复 Windows 终端 GBK 编码问题（支持 emoji 和中文输出）
if sys.stdout.encoding != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# 确保项目根目录在 sys.path 中（支持直接 python src/main.py 执行）
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from src.config_loader import (
    ConfigError,
    get_default_config_path,
    list_categories,
    load_config,
    resolve_category_config,
)
from src.processors.base_processor import (
    BaseProcessor,
    IOFailure,
    ProcessingError,
)

logger = logging.getLogger(__name__)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Amazon 商品数据 → 模板填充处理器（配置驱动）",
    )
    sub = parser.add_subparsers(dest="command", help="可用子命令")

    # ---- process 子命令 ----
    proc = sub.add_parser("process", help="执行数据处理流水线")
    proc.add_argument(
        "--type",
        required=True,
        help="产品类型标识（如 azd, hd, th）。必选。",
    )
    proc.add_argument(
        "--source",
        default=None,
        help="源 Excel 文件路径（覆盖配置文件中的默认值）",
    )
    proc.add_argument(
        "--template",
        default=None,
        help="模板文件路径（覆盖配置文件中的默认值）",
    )
    proc.add_argument(
        "--chunk",
        type=int,
        default=None,
        help="分块行数（覆盖配置文件中的默认值）",
    )
    proc.add_argument(
        "--output-dir",
        default=None,
        help="输出目录（覆盖配置文件中的默认值）",
    )
    proc.add_argument(
        "--config",
        default=None,
        help="配置文件路径（默认: config/categories.yaml）",
    )
    proc.add_argument(
        "--dry-run",
        action="store_true",
        help="只做预处理分析，不写文件",
    )

    # ---- list-types 子命令 ----
    lst = sub.add_parser("list-types", help="列出所有可用产品类型")
    lst.add_argument(
        "--config",
        default=None,
        help="配置文件路径（默认: config/categories.yaml）",
    )

    # ---- price 子命令 ----
    price = sub.add_parser("price", help="利润定价：根据成本+重量反算销售价并回填")
    price.add_argument(
        "--source",
        required=True,
        help="源 Excel 文件路径（必选，结果将回填到该文件）。",
    )
    price.add_argument(
        "--target",
        type=float,
        default=0.25,
        help="目标毛利率（默认: 0.25 即 25%%）。",
    )
    price.add_argument(
        "--rate",
        type=float,
        default=7.8,
        help="汇率 RMB/外币（默认: 7.8）。",
    )
    price.add_argument(
        "--shipping-price",
        type=float,
        default=0.0,
        help="向买家收取的运费-外币（默认: 0，包邮模式）。",
    )
    price.add_argument(
        "--shipping-method",
        choices=["ceil", "linear", "floor"],
        default="ceil",
        help="运费查表方式（默认: ceil 向上找档）。",
    )
    price.add_argument(
        "--station",
        choices=["DE", "US", "CA"],
        default="DE",
        help="目标站点（默认: DE）。",
    )
    price.add_argument(
        "--cost-col",
        default="I",
        help="成本价(RMB)所在列字母（默认: I）。",
    )
    price.add_argument(
        "--weight-col",
        default="J",
        help="重量(g)所在列字母（默认: J）。",
    )
    price.add_argument(
        "--price-col",
        default="AM",
        help="销售价(EUR)写入列字母（默认: AM）。",
    )

    # ---- translate 子命令 ----
    translate = sub.add_parser(
        "translate",
        help="颜色翻译与变体检测（英→德翻译 + V1/V2 后缀，原地修改源文件）",
    )
    translate.add_argument(
        "--source",
        required=True,
        help="源 Excel 文件路径（必选，结果将原地回填）。",
    )
    translate.add_argument(
        "--color-col",
        default="G",
        help="颜色所在列字母（默认: G）。",
    )
    translate.add_argument(
        "--size-col",
        default="H",
        help="尺码所在列字母（默认: H）。",
    )
    translate.add_argument(
        "--sheet",
        type=int,
        default=0,
        help="工作表索引，从 0 开始（默认: 0 即第一个工作表）。",
    )
    translate.add_argument(
        "--no-translate",
        action="store_true",
        help="跳过颜色翻译，仅执行变体检测。",
    )
    translate.add_argument(
        "--no-variant",
        action="store_true",
        help="跳过变体检测，仅执行颜色翻译。",
    )

    return parser


def cmd_process(args: argparse.Namespace) -> None:
    """执行 process 子命令。"""
    # 1. 加载配置文件
    config_path = args.config or get_default_config_path()
    print(f"📋 配置文件: {config_path}")
    config = load_config(config_path)

    # 2. 解析产品类型配置
    cat_config = resolve_category_config(config, args.type)
    description = cat_config.get("description", args.type)
    print(f"🏷️  产品类型: {args.type} ({description})")

    # 3. CLI 参数覆盖配置文件默认值
    if args.source:
        cat_config["default_source"] = args.source
    if args.template:
        cat_config["template_file"] = args.template
    if args.chunk:
        cat_config["chunk_size"] = args.chunk
    if args.output_dir:
        cat_config["output_dir"] = args.output_dir

    # 4. 创建处理器并执行
    processor = BaseProcessor(cat_config)

    if args.dry_run:
        processor.dry_run()
    else:
        processor.run()


def cmd_price(args: argparse.Namespace) -> None:
    """执行 price 子命令：根据成本+重量反算销售价并回填到源文件。"""
    from src.profit_calculator import ProfitCalculator, Station

    source = args.source
    if not os.path.exists(source):
        raise FileNotFoundError(f"源文件不存在: {source}")

    station = Station(args.station)
    cost_idx = column_index_from_string(args.cost_col) - 1    # 0-based
    weight_idx = column_index_from_string(args.weight_col) - 1
    price_idx = column_index_from_string(args.price_col) - 1

    print(f"📋 源文件:   {source}")
    print(f"🏷️  站点:     {station.value}")
    print(f"🎯 目标毛利: {args.target:.1%}")
    print(f"💱 汇率:     {args.rate}")
    print(f"📦 运费查表: {args.shipping_method}")
    print(f"📐 成本列:   {args.cost_col}  重量列: {args.weight_col}  →  写入列: {args.price_col}")
    print()

    # 1. 用 pandas 读取数据
    df = pd.read_excel(source, header=None)

    # 确保目标列存在
    while df.shape[1] <= price_idx:
        df[df.shape[1]] = np.nan

    # 2. 创建计算器
    calc = ProfitCalculator(station=station)

    success = 0
    skipped = 0
    price_list = []

    for idx in range(len(df)):
        cost = df.iloc[idx, cost_idx]
        weight = df.iloc[idx, weight_idx]

        if pd.isna(cost) or pd.isna(weight):
            skipped += 1
            continue
        try:
            cost_val = float(cost)
            weight_val = float(weight)
        except (ValueError, TypeError):
            skipped += 1
            continue
        if cost_val <= 0 or weight_val <= 0:
            skipped += 1
            continue

        try:
            sell_price = calc.breakeven_price(
                exchange_rate=args.rate,
                sku_cost_rmb=cost_val,
                weight_g=weight_val,
                shipping_price=args.shipping_price,
                target_margin=args.target,
                shipping_method=args.shipping_method,
            )
        except ValueError as e:
            logger.warning("Row %d: %s", idx, e)
            skipped += 1
            continue

        sell_price = round(sell_price, 2)
        df.iloc[idx, price_idx] = sell_price
        price_list.append(sell_price)
        success += 1

    if success == 0:
        print("⚠️  没有找到有效数据行（成本/重量为空或 ≤0）。")
        return

    # 3. 统计
    prices = pd.Series(price_list)
    print(f"✅ 有效行数: {success}  ({skipped} 行跳过)")
    print(f"💰 价格区间: EUR {prices.min():.2f} ~ {prices.max():.2f}")
    print(f"📊 价格均值: EUR {prices.mean():.2f}  中位数: EUR {prices.median():.2f}")

    # 4. 用 openpyxl 回填到源文件（仅写入售价列，保留其它列不变）
    print(f"\n💾 回填到源文件: {source}")
    wb = load_workbook(source)
    ws = wb.active

    # 写入表头
    ws.cell(row=1, column=price_idx + 1, value="销售价(EUR)")

    # 写入数据（Excel 行号 = DataFrame idx + 2，因为 Excel 第 1 行留作表头）
    for idx in range(len(df)):
        val = df.iloc[idx, price_idx]
        if pd.notna(val):
            ws.cell(row=idx + 2, column=price_idx + 1, value=float(val))

    wb.save(source)
    wb.close()

    print(f"🎉 已将 {success} 条销售价写入 {args.price_col} 列。")


def cmd_translate(args: argparse.Namespace) -> None:
    """执行 translate 子命令：颜色翻译 + 变体检测，原地修改源文件。"""
    from src.color_translator import ColorTranslator

    source = args.source
    if not os.path.exists(source):
        raise FileNotFoundError(f"源文件不存在: {source}")

    color_idx = column_index_from_string(args.color_col) - 1
    size_idx = column_index_from_string(args.size_col) - 1

    print(f"📋 源文件:   {source}")
    print(f"🎨 颜色列:   {args.color_col} (0-based 索引 {color_idx})")
    print(f"📏 尺码列:   {args.size_col} (0-based 索引 {size_idx})")
    print(f"📄 工作表:   第 {args.sheet} 个")
    print(f"🌐 颜色翻译: {'❌ 关' if args.no_translate else '✅ 开'}")
    print(f"🔢 变体检测: {'❌ 关' if args.no_variant else '✅ 开'}")
    print()

    # 1. 用 pandas 读取数据
    df = pd.read_excel(source, sheet_name=args.sheet, header=None)

    # 2. 处理颜色翻译和变体检测
    ct = ColorTranslator()
    result = ct.process(
        df,
        color_col=color_idx,
        size_col=size_idx,
        parent_sku_col=2,
        enable_translation=not args.no_translate,
        enable_variant=not args.no_variant,
    )

    # 3. 统计
    changed = 0
    for idx, new_color in result.index_to_new_color.items():
        old = df.iloc[idx, color_idx]
        old_str = str(old).strip() if pd.notna(old) else ""
        if old_str != new_color:
            changed += 1

    print(f"📊 总处理行数:   {result.total_rows}")
    print(f"📊 颜色翻译数:   {result.total_translated} 个颜色组")
    print(f"📊 变体标注数:   {result.total_variant_labeled} 行")
    print(f"📊 实际修改行数: {changed}")
    print()

    if changed == 0:
        print("✨ 没有需要修改的行。")
        return

    # 4. 用 openpyxl 原地回填颜色列
    print(f"💾 回填到源文件: {source}")
    wb = load_workbook(source)
    ws = wb.worksheets[args.sheet]

    for idx, new_color in result.index_to_new_color.items():
        old = df.iloc[idx, color_idx]
        old_str = str(old).strip() if pd.notna(old) else ""
        if old_str != new_color:
            # Excel 行号 = DataFrame idx + 1（无表头，数据从第 1 行开始）
            ws.cell(row=idx + 1, column=color_idx + 1, value=new_color)

    wb.save(source)
    wb.close()

    print(f"🎉 已将 {changed} 行颜色值写回 {args.color_col} 列。")


def cmd_list_types(args: argparse.Namespace) -> None:
    """执行 list-types 子命令。"""
    config_path = args.config or get_default_config_path()
    print(f"📋 配置文件: {config_path}")
    config = load_config(config_path)

    categories = list_categories(config)
    if not categories:
        print("（无可用产品类型）")
        return

    print(f"\n{'标识':<8} 描述")
    print("-" * 50)
    for cat in categories:
        print(f"{cat['key']:<8} {cat['description']}")


def main(argv: Optional[list[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.command:
        parser.print_help()
        sys.exit(0)

    try:
        if args.command == "process":
            cmd_process(args)
        elif args.command == "list-types":
            cmd_list_types(args)
        elif args.command == "price":
            cmd_price(args)
        elif args.command == "translate":
            cmd_translate(args)
    except ConfigError as e:
        print(f"❌ 配置错误: {e}", file=sys.stderr)
        sys.exit(1)
    except ProcessingError as e:
        print(f"❌ 处理错误: {e}", file=sys.stderr)
        sys.exit(1)
    except IOFailure as e:
        print(f"❌ 文件错误: {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError as e:
        print(f"❌ 文件未找到: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"❌ 未预期的错误: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()

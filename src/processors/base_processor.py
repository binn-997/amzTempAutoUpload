"""
品牌处理器基类 — 配置驱动，所有产品类型通用。

接收一个配置字典（由 config_loader 解析 YAML/JSON 后提供），
执行完整的数据处理流水线：读取 → 预处理 → 填充 → 清除 → 保存。

用法:
    from src.config_loader import load_config, resolve_category_config
    from src.processors.base_processor import BaseProcessor

    config = load_config("config/categories.yaml")
    cat_config = resolve_category_config(config, "azd")
    # CLI 参数可在此覆盖 cat_config 中的值
    processor = BaseProcessor(cat_config)
    processor.run()
"""

from __future__ import annotations

import io
import logging
import math
import os
import re
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ==================== 异常体系 ====================


class ProcessingError(Exception):
    """处理流程中的可恢复错误。"""


class IOFailure(RuntimeError):
    """Excel 文件读写失败（携带文件名和操作类型）。"""


# ==================== 工具函数 ====================


def sanitize_cell_value(value: Any) -> Any:
    """检测并转义潜在的 CSV/Excel 公式注入字符。

    若字符串以 = + - @ 开头，前缀单引号中和公式执行。
    模块级函数，可供 process.py / split_excel.py 等外部工具复用。
    """
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped and stripped[0] in ("=", "+", "-", "@"):
            logger.warning(
                "检测到潜在公式注入，已转义 [%s...]", str(value)[:80]
            )
            return "'" + value
    return value


# ==================== BaseProcessor ====================


class BaseProcessor:
    """Amazon 商品数据 → 模板填充处理器的配置驱动实现。

    所有产品类型的差异化参数均来源于构造函数传入的 config 字典。
    类属性仅作为回退默认值（当 config 中未提供时使用）。
    """

    # ==================== 构造函数 ====================

    def __init__(self, config: dict | None = None):
        """初始化处理器。

        Args:
            config: 产品类型配置字典（由 config_loader.resolve_category_config() 提供）。
                    应为已合并 defaults 的完整配置。
                    若为 None 或空，则使用类属性回退值（向后兼容）。
        """
        self._config = config or {}

        # ── 从 config 提取所有参数 ──
        self.source_file: str = self._get("default_source", "")
        self.template_file: str = self._get("template_file", "")
        self.chunk_size: int = int(self._get("chunk_size", 500))
        self.title_prefix: str = self._get("title_prefix", "")
        self.output_dir: str = self._get("output_dir", "./prov_output")
        self.output_suffix: str = self._get("output_suffix", "_prov")
        self.include_final_keywords: bool = bool(
            self._get("include_final_keywords", True)
        )
        self.size_class_col: Optional[str] = self._get("size_class_col", "AD")
        self.size_target_cols: List[str] = list(
            self._get("size_target_cols", ["AE", "CH", "FC"])
        )
        self.parent_clear_letters: List[str] = list(
            self._get("parent_clear_letters", [
                "V", "W", "X", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                "BY", "BZ", "CF", "CG", "CH", "FC",
            ])
        )
        self.size_replacements: Dict[str, str] = dict(
            self._get("size_replacements", {
                r"\bXXXL\b": "3XL",
                r"\bXXXXL\b": "4XL",
                r"\bXXXXXL\b": "5XL",
                r"\bXXXXXXL\b": "6XL",
                r"\bXXXXXXXL\b": "7XL",
                r"\bXXXXXXXXL\b": "8XL",
                r"\bXXXXXXXXXL\b": "9XL",
                r"\bone size\b": "Einheitsgröße",
            })
        )

        # ── 共享的源列映射（默认值 + config 全量透传，支持自定义字段） ──
        src_cols = self._config.get("source_columns", {})
        _src_defaults: Dict[str, Any] = {
            "sku": 1,
            "parent_sku": 2,
            "title": 3,
            "color": 6,
            "size": 7,
            "package_weight": 9,
            "standprice": 38,
            "list_price_tax": 39,
            "generic_keywords": 45,
            "bullet1": 40,
            "bullet2": 41,
            "bullet3": 42,
            "bullet4": 43,
            "bullet5": 44,
            "images": [10, 11, 12, 13, 14, 15, 16],
        }
        # config 中的值覆盖默认值，config 中的新键（如 Produktbeschreibungen）直接添加
        self.src = {**_src_defaults, **src_cols}

        # ── 共享的模板列映射（来自 config，回退到类常量） ──
        tpl_cols = self._config.get("template_columns", {})
        tpl_simple = tpl_cols.get("simple", {})
        tpl_multi = tpl_cols.get("multi", {})
        tpl_bullets = tpl_cols.get("bullets", ["CO", "CP", "CQ", "CR", "CS"])

        _tpl_simple_defaults: Dict[str, str] = {
            "sku": "B",
            "parent_sku": "BY",
            "title": "G",
            "standprice": "V",
            "package_weight": "GD",
            "list_price_tax": "KP",
            "final_keywords": "CE",
        }
        # config 中的值覆盖默认值，config 中的新键（如 Produktbeschreibungen）直接添加
        self._tpl_simple_base = {**_tpl_simple_defaults, **tpl_simple}
        self._tpl_multi_color: List[str] = list(
            tpl_multi.get("color", ["CF", "CG"])
        )
        self._tpl_bullets: List[str] = list(tpl_bullets)
        self._tpl_images_start: str = tpl_cols.get("images_start", "BH")
        self._tpl_pc_col: str = tpl_cols.get("parent_child", "BX")

        # ── 其他常量 ──
        self.max_source_cols: int = 60
        self.start_row: int = 4
        self.max_images: int = 7
        self.sheet_name: str = "Vorlage"
        self.bullet_src_prefix: str = "final_bullet"
        self.img_src_cols: List[int] = [10, 11, 12, 13, 14, 15, 16]

        # ── 可选的预处理 / 填充钩子（默认无操作） ──
        self._preprocess_hooks: List[Callable[[pd.DataFrame], pd.DataFrame]] = []
        self._post_fill_hooks: List[Callable[[Worksheet, int, pd.Series], None]] = []

    def _get(self, key: str, default: Any = None) -> Any:
        """从配置字典安全获取值，key 不存在时返回 default。"""
        return self._config.get(key, default)

    # ==================== 计算属性 ====================

    @property
    def parent_clear_idx(self) -> List[int]:
        """父体清空列的 1-based 索引列表（延迟计算）。"""
        return [column_index_from_string(c) for c in self.parent_clear_letters]

    @property
    def simple_map(self) -> Dict[str, str]:
        """一对一列映射：{源列名: 目标 Excel 列字母}。"""
        m: Dict[str, str] = dict(self._tpl_simple_base)
        if self.size_class_col:
            m["size_class"] = self.size_class_col
        if not self.include_final_keywords:
            m.pop("final_keywords", None)
        return m

    @property
    def multi_map(self) -> Dict[str, List[str]]:
        """一对多列映射：{源列名: [目标 Excel 列字母列表]}。"""
        return {
            "color": list(self._tpl_multi_color),
            "size": list(self.size_target_cols),
        }

    @property
    def _simple_idx(self) -> Dict[str, int]:
        return {k: column_index_from_string(v) for k, v in self.simple_map.items()}

    @property
    def _multi_idx(self) -> Dict[str, List[int]]:
        return {
            k: [column_index_from_string(c) for c in v]
            for k, v in self.multi_map.items()
        }

    @property
    def _bullet_idx(self) -> List[int]:
        return [column_index_from_string(c) for c in self._tpl_bullets]

    @property
    def _img_start(self) -> int:
        return int(column_index_from_string(self._tpl_images_start))

    @property
    def _pc_idx(self) -> int:
        return int(column_index_from_string(self._tpl_pc_col))

    # ==================== 钩子方法（子类/外部可覆盖） ====================

    def post_preprocess(self, df: pd.DataFrame) -> pd.DataFrame:
        """预处理后钩子 — 产品类型特定的数据清洗逻辑。

        在 preprocess_source_data 的最后一步调用。
        默认实现为空操作（返回原 df）。

        如需添加产品类型特定逻辑，可在配置中指定 hook 名称，
        然后在子类或外部覆盖此方法。
        """
        return df

    def post_fill_row(
        self, ws: Worksheet, excel_row: int, row_data: pd.Series
    ) -> None:
        """单行填充后钩子 — 产品类型特定的单元格后处理。

        在 fill_data_to_template 中对每一行调用。
        默认实现为空操作。

        例如：某些产品类型可能需要在特定列写入额外公式或格式。
        """
        pass

    # ==================== 数据预处理 ====================

    def preprocess_source_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """在内存中完成所有数据清洗与父子体继承。"""
        print("预处理源数据：清洗空值、替换尺码、继承父体数据...")

        # 列补齐 + 截断 + 空值填充
        if df.shape[1] < self.max_source_cols:
            for i in range(df.shape[1], self.max_source_cols):
                df[i] = ""
        df = df.iloc[:, : self.max_source_cols].fillna("")
        df = df.apply(
            lambda col: col.str.strip() if col.dtype == "object" else col
        )

        # 列重命名（基于 src 映射）
        rename_map = {v: k for k, v in self.src.items() if isinstance(v, int)}
        df = df.rename(columns=rename_map)

        # title 加前缀
        if self.title_prefix:
            df["title"] = self.title_prefix + df["title"].astype(str)

        # 尺码替换
        for pattern, replacement in self.size_replacements.items():
            df["size"] = (
                df["size"]
                .astype(str)
                .str.replace(pattern, replacement, regex=True, flags=re.IGNORECASE)
            )

        # size_class
        if self.size_class_col:
            df["size_class"] = df["size"].apply(
                lambda x: (
                    "Numerisch" if str(x).strip().isdigit() else "Alphanumerisch"
                )
            )

        # 父子标识
        df["is_parent"] = df["sku"] == df["parent_sku"]

        # 子体从父体继承 keywords & bullets
        parent_cols = [
            "generic_keywords",
            "bullet1", "bullet2", "bullet3", "bullet4", "bullet5",
        ]
        parent_df = (
            df[df["is_parent"]][["sku"] + parent_cols].set_index("sku")
        )
        parent_dict: dict = parent_df.to_dict(orient="index")

        def get_parent_data(row: pd.Series) -> pd.Series:
            if row["is_parent"]:
                return pd.Series([row[col] for col in parent_cols])
            p_data = parent_dict.get(row["parent_sku"], {})
            return pd.Series([p_data.get(col, "") for col in parent_cols])

        df[
            ["final_keywords"] + [f"{self.bullet_src_prefix}{i}" for i in range(1, 6)]
        ] = df.apply(get_parent_data, axis=1)

        # 执行预处理钩子
        df = self.post_preprocess(df)

        return df

    # ==================== 模板填充 ====================

    def fill_data_to_template(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """根据映射配置，将 DataFrame 逐行写入模板工作表。"""
        for r_idx, (_, row) in enumerate(df.iterrows(), start=self.start_row):
            # A. 一对一映射
            for src_col, tgt_idx in self._simple_idx.items():
                ws.cell(
                    row=r_idx,
                    column=tgt_idx,
                    value=sanitize_cell_value(row[src_col]),
                )

            # B. 一对多映射
            for src_col, tgt_indices in self._multi_idx.items():
                for tgt_idx in tgt_indices:
                    ws.cell(
                        row=r_idx,
                        column=tgt_idx,
                        value=sanitize_cell_value(row[src_col]),
                    )

            # C. Bullet Points
            for i, tgt_idx in enumerate(self._bullet_idx):
                src_col = f"{self.bullet_src_prefix}{i + 1}"
                ws.cell(
                    row=r_idx,
                    column=tgt_idx,
                    value=sanitize_cell_value(row[src_col]),
                )

            # D. 图片
            images = [
                str(row[pos]).strip()
                for pos in self.img_src_cols
                if str(row[pos]).strip()
            ]
            for img_i, url in enumerate(images[: self.max_images]):
                ws.cell(
                    row=r_idx,
                    column=self._img_start + img_i,
                    value=sanitize_cell_value(url),
                )
            for empty_i in range(len(images), self.max_images):
                ws.cell(row=r_idx, column=self._img_start + empty_i, value="")

            # E. 父子标识
            pc_val = "Parent" if row["is_parent"] else "Child"
            ws.cell(row=r_idx, column=self._pc_idx, value=pc_val)

            # F. 单行后处理钩子
            self.post_fill_row(ws, r_idx, row)

    # ==================== 父体清除 ====================

    def clear_parent_rows(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """抹平父体行的子体专属字段，并取消合并单元格。"""
        parent_local_indices = df[df["is_parent"]].index.tolist()
        if not parent_local_indices:
            return

        for local_row_idx in parent_local_indices:
            excel_row = self.start_row + local_row_idx

            # 取消该行所有合并单元格
            merged_to_remove = [
                mr
                for mr in ws.merged_cells.ranges
                if mr.min_row <= excel_row <= mr.max_row
            ]
            for mr in merged_to_remove:
                ws.unmerge_cells(str(mr))

            # 清空父体专属字段
            for col_idx in self.parent_clear_idx:
                ws.cell(row=excel_row, column=col_idx, value="")

            ws.cell(row=excel_row, column=self._pc_idx, value="Parent")

    # ==================== 拆分保存 ====================

    def _load_template_buffer(self) -> io.BytesIO:
        """加载模板文件到内存缓冲区（只读一次磁盘）。"""
        try:
            wb = load_workbook(self.template_file, keep_vba=False)
        except Exception as e:
            raise IOFailure(
                f"加载模板失败 [{self.template_file}]: {e}"
            ) from e

        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ProcessingError(
                f"模板中未找到 '{self.sheet_name}' 工作表"
            )

        buf = io.BytesIO()
        try:
            wb.save(buf)
        except Exception as e:
            raise IOFailure(
                f"缓存模板失败 [{self.template_file}]: {e}"
            ) from e
        finally:
            wb.close()
        return buf

    def _create_chunk_workbook(
        self, template_buf: io.BytesIO
    ) -> tuple[Workbook, Worksheet]:
        """从内存缓冲区创建分块工作簿，返回 (workbook, worksheet)。"""
        template_buf.seek(0)
        try:
            wb = load_workbook(template_buf, keep_vba=False)
        except Exception as e:
            raise IOFailure(f"从缓存创建分块失败: {e}") from e
        return wb, wb[self.sheet_name]

    def save_split_workbooks(self, processed_df: pd.DataFrame) -> None:
        """将数据按 chunk_size 拆分为多个文件并保存。

        优化：模板文件只在开始时加载一次到内存 BytesIO，
        后续分块均从内存缓存创建，避免重复磁盘 I/O。
        """
        total_rows = len(processed_df)
        if total_rows == 0:
            print("源数据为空，跳过。")
            return

        # 准备输出目录
        output_dir = self.output_dir
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"📁 创建输出目录: {output_dir}")

        base_name = os.path.splitext(os.path.basename(self.source_file))[0]
        num_chunks = math.ceil(total_rows / self.chunk_size)

        print(f"\n数据预处理完成，共 {total_rows} 行。")
        print(
            f"按每文件 {self.chunk_size} 行拆分，共 {num_chunks} 个文件。\n"
        )

        # ★ 性能优化：模板只加载一次到内存
        print(f"📖 加载模板: {self.template_file}")
        template_buf = self._load_template_buffer()

        for chunk_idx in range(num_chunks):
            start = chunk_idx * self.chunk_size
            end = min(start + self.chunk_size, total_rows)

            chunk_df = processed_df.iloc[start:end].reset_index(drop=True)
            chunk_rows = len(chunk_df)

            wb, ws = self._create_chunk_workbook(template_buf)

            # 填充 & 清除
            self.fill_data_to_template(ws, chunk_df)
            self.clear_parent_rows(ws, chunk_df)

            # 删除多余空行
            last_used_row = self.start_row + chunk_rows - 1
            start_delete_row = last_used_row + 1
            if ws.max_row >= start_delete_row:
                ws.delete_rows(
                    start_delete_row,
                    amount=ws.max_row - start_delete_row + 1,
                )

            output_name = (
                f"{base_name}{self.output_suffix}_part{chunk_idx + 1}.xlsx"
            )
            if output_dir:
                output_path = os.path.join(output_dir, output_name)
            else:
                output_path = output_name

            print(
                f"  [{chunk_idx + 1}/{num_chunks}] 💾 {output_path}"
                f"  ({chunk_rows} 行)"
            )
            try:
                wb.save(output_path)
            except Exception as e:
                wb.close()
                raise IOFailure(
                    f"保存文件失败 [{output_path}]: {e}"
                ) from e
            wb.close()

        print(f"\n✅ 全部 {num_chunks} 个文件保存完毕。")

    # ==================== 入口 ====================

    def _load_and_preprocess(self) -> pd.DataFrame:
        """读取源 Excel 文件并执行预处理（run / dry_run 共用）。"""
        if not self.source_file:
            raise ProcessingError("SOURCE_FILE 未设置")
        if not self.template_file:
            raise ProcessingError("TEMPLATE_FILE 未设置")

        if not os.path.exists(self.source_file):
            raise ProcessingError(f"源文件不存在: {self.source_file}")
        if not os.path.exists(self.template_file):
            raise ProcessingError(f"模板文件不存在: {self.template_file}")

        try:
            raw_df: pd.DataFrame = pd.read_excel(
                self.source_file, sheet_name=0, header=None, dtype=str
            )
        except Exception as e:
            raise IOFailure(
                f"读取源文件失败 [{self.source_file}]: {e}"
            ) from e

        return self.preprocess_source_data(raw_df)

    def run(self) -> None:
        """执行完整处理流水线：读取 → 预处理 → 拆分写入。"""
        print(f"🚀 开始执行 Amazon 模板处理程序 [{self.__class__.__name__}]")

        print(f"\n[1/3] 读取源数据: {self.source_file}")
        print("\n[2/3] 预处理源数据...")
        processed_df = self._load_and_preprocess()

        print("\n[3/3] 执行拆分写入...")
        self.save_split_workbooks(processed_df)

        print("\n🎉 全部流程执行完毕！")

    def dry_run(self) -> None:
        """只做预处理，不写文件 — 用于验证配置。"""
        print(f"[DRY-RUN] 模式 [{self.__class__.__name__}]")

        processed_df = self._load_and_preprocess()

        total_rows = len(processed_df)
        num_chunks = math.ceil(total_rows / self.chunk_size)
        parent_count = processed_df["is_parent"].sum()
        child_count = total_rows - parent_count

        print(f"\n[SUMMARY] 预处理结果:")
        print(f"   总行数:     {total_rows}")
        print(f"   父体:       {parent_count}")
        print(f"   子体:       {child_count}")
        print(f"   拆分文件数: {num_chunks} (每 {self.chunk_size} 行)")
        print(f"   列数:       {len(processed_df.columns)}")
        print(f"   输出目录:   {self.output_dir}")
        print(f"   模板文件:   {self.template_file}")
        print(f"\n[DONE] Dry-run 完成（未写入任何文件）。")

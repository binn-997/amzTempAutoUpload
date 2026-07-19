from openpyxl import Workbook, load_workbook

from amazon_listing_toolkit.excel_splitter import split_workbook


def test_split_workbook_preserves_header_and_chunks(tmp_path) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vorlage"
    sheet.append(["header"])
    sheet.append(["header"])
    sheet.append(["header"])
    for value in range(5):
        sheet.append([value])
    workbook.save(source)
    workbook.close()

    split_workbook(str(source), chunk_size=2, output_dir=str(tmp_path / "out"))

    first = load_workbook(tmp_path / "out" / "source_part1.xlsx")
    assert first["Vorlage"].max_row == 5
    first.close()
    assert (tmp_path / "out" / "source_part3.xlsx").exists()

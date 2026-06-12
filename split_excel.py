#!/usr/bin/env python3
"""
============================================================
 纯拆分工具 — 支持拖放 Excel 直接拆分
============================================================
用法：
  ① 把 Excel 拖到 exe 图标上          → 直接拆分该文件
  ② 双击 exe                          → 自动检测当前目录 Excel
  ③ 命令行: 拆分表工具.exe -s 文件.xlsx -n 300

发给别人：把 .exe 放到任意文件夹，拖 Excel 上去即可。
"""

import os
import sys
import math
import argparse
from pathlib import Path
from openpyxl import load_workbook


# ============================================================
#  默  认  参  数  （只设行为，不设文件名）
# ============================================================
CHUNK_SIZE    = 500
HEADER_ROWS   = 3
SHEET_NAME    = "Vorlage"
OUTPUT_DIR    = "./split_output"
OUTPUT_SUFFIX = "_part"


# ============================================================
#  自  动  检  测
# ============================================================

def find_excel_files(directory: str = ".") -> list:
    extensions = {".xlsx", ".xlsm"}
    files = []
    for f in Path(directory).iterdir():
        if f.suffix.lower() in extensions and not f.name.startswith("~$"):
            files.append(str(f))
    return sorted(files)


def pick_source_file() -> str:
    candidates = find_excel_files()
    if not candidates:
        print("❌ 当前目录没有找到 .xlsx / .xlsm 文件。")
        print("   请把要拆分的 Excel 拖到本程序图标上，或放到同目录下。")
        input("\n按回车键退出...")
        sys.exit(1)

    if len(candidates) == 1:
        print(f"🔍 自动检测到: {candidates[0]}")
        return candidates[0]

    print("\n📂 当前目录有多个 Excel，请选择：")
    for i, f in enumerate(candidates, 1):
        print(f"   [{i}] {Path(f).name}")
    while True:
        try:
            choice = input(f"   输入序号 (1-{len(candidates)}): ").strip()
            idx = int(choice) - 1
            if 0 <= idx < len(candidates):
                return candidates[idx]
        except ValueError:
            pass
        print(f"   ⚠️ 请输入 1-{len(candidates)}")


# ============================================================
#  核  心  拆  分
# ============================================================

def split_workbook(
    source_path: str,
    chunk_size: int = CHUNK_SIZE,
    header_rows: int = HEADER_ROWS,
    sheet_name: str = SHEET_NAME,
    output_dir: str = OUTPUT_DIR,
    output_suffix: str = OUTPUT_SUFFIX,
):
    print(f"📖 加载: {source_path}")
    wb = load_workbook(source_path, keep_vba=True)

    actual_sheet = sheet_name
    if actual_sheet not in wb.sheetnames:
        visible = [s for s in wb.sheetnames if wb[s].sheet_state == 'visible']
        actual_sheet = visible[0] if visible else wb.sheetnames[0]
        print(f"   工作表 '{sheet_name}' 不存在 → 改用 '{actual_sheet}'")

    ws = wb[actual_sheet]
    max_row = ws.max_row
    data_start = header_rows + 1
    total_data_rows = max_row - header_rows

    if total_data_rows <= 0:
        print("⚠️  没有数据行，无需拆分。")
        wb.close()
        return

    num_chunks = math.ceil(total_data_rows / chunk_size)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"📁 输出目录: {output_dir}")

    base_name = Path(source_path).stem

    print(f"   总行数: {max_row}  |  表头: {header_rows} 行  |  数据: {total_data_rows} 行")
    print(f"   每文件 ≤ {chunk_size} 行  →  {num_chunks} 个文件\n")

    for chunk_idx in range(num_chunks):
        chunk_start = data_start + chunk_idx * chunk_size
        chunk_end   = min(data_start + (chunk_idx + 1) * chunk_size - 1, max_row)
        chunk_rows  = chunk_end - chunk_start + 1

        new_wb = load_workbook(source_path, keep_vba=True)
        new_ws = new_wb[actual_sheet]

        if max_row > chunk_end:
            new_ws.delete_rows(chunk_end + 1, amount=max_row - chunk_end)
        if chunk_start > data_start:
            new_ws.delete_rows(data_start, amount=chunk_start - data_start)

        output_name = f"{base_name}{output_suffix}{chunk_idx + 1}.xlsx"
        output_path = os.path.join(output_dir, output_name)

        new_ws.sheet_state = 'visible'
        print(f"  [{chunk_idx + 1}/{num_chunks}] 💾 {output_name}  ({chunk_rows} 行)")
        new_wb.save(output_path)
        new_wb.close()

    wb.close()
    print(f"\n🎉 完成！{num_chunks} 个文件 → {output_dir}")


# ============================================================
#  入  口
# ============================================================
def main():
    print("=" * 50)
    print("  ✂️  纯拆分工具 — 拖放 Excel 即可拆分")
    print("=" * 50)

    # ── 1. 拖放优先：如果 sys.argv 有裸路径（非 -x 开头），直接当源文件 ──
    #    拖放时 Windows 传的是: "程序.exe" "C:\Users\...\文件.xlsx"
    raw_args = [a for a in sys.argv[1:] if not a.startswith("-")]

    if raw_args:
        # 拖放模式
        source = raw_args[0]
        if not os.path.exists(source):
            print(f"❌ 文件不存在: {source}")
            input("\n按回车键退出...")
            sys.exit(1)
        print(f"🖱️  拖放文件: {Path(source).name}")

        # 剩余命令行参数仍可用 -n -o 等
        parser = argparse.ArgumentParser()
        parser.add_argument("source", nargs="?", default=source)
        parser.add_argument("-n", "--chunk-size", type=int, default=CHUNK_SIZE)
        parser.add_argument("-o", "--output-dir", default=OUTPUT_DIR)
        parser.add_argument("--header-rows", type=int, default=HEADER_ROWS)
        parser.add_argument("--sheet", default=SHEET_NAME)
        parser.add_argument("--suffix", default=OUTPUT_SUFFIX)
        # 用 parse_known_args 避免裸路径被当成未知参数
        args, _ = parser.parse_known_args()
    else:
        # ── 2. 没有拖放：正常 argparse ──
        parser = argparse.ArgumentParser(description="纯拆分工具 — 拖放 Excel 到 exe 即可拆分")
        parser.add_argument("-s", "--source", default=None, help="源 Excel（不填则自动检测）")
        parser.add_argument("-n", "--chunk-size", type=int, default=CHUNK_SIZE)
        parser.add_argument("-o", "--output-dir", default=OUTPUT_DIR)
        parser.add_argument("--header-rows", type=int, default=HEADER_ROWS)
        parser.add_argument("--sheet", default=SHEET_NAME)
        parser.add_argument("--suffix", default=OUTPUT_SUFFIX)
        args = parser.parse_args()

        source = args.source or pick_source_file()

    print(f"📐 每文件上限: {args.chunk_size} 行")
    print(f"📏 表头行数:   {args.header_rows}")
    print(f"📋 工作表:     {args.sheet}\n")

    split_workbook(
        source_path=source,
        chunk_size=args.chunk_size,
        header_rows=args.header_rows,
        sheet_name=args.sheet,
        output_dir=args.output_dir,
        output_suffix=args.suffix,
    )

    # 拖放 / 双击模式下防止窗口闪退
    if not raw_args or len(sys.argv) <= 2:
        input("\n按回车键退出...")


if __name__ == "__main__":
    main()